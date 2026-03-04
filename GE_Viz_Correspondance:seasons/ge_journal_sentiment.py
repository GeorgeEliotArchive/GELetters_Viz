"""
George Eliot Journal Sentiment Analysis
========================================
Extracts all journal entries written BY George Eliot from the Excel database
and applies two lexicons appropriate for 19th-century literary prose:

  1. SentiWordNet  — built on WordNet's formal dictionary synsets; covers
                     literary, archaic, and formal vocabulary far better than
                     social-media lexicons like VADER.

  2. NRC Emotion Lexicon — hand-annotated by linguists for ~14,000 words
                     across 8 emotions + positive/negative polarity; designed
                     for broad-domain text including literature.

"""

import re
import warnings
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from openpyxl import load_workbook

import nltk
from nltk.corpus import sentiwordnet as swn
from nltk.tokenize import word_tokenize
from nltk import pos_tag
from nrclex import NRCLex
from sklearn.preprocessing import MinMaxScaler

warnings.filterwarnings("ignore")

# ── 0. Download required NLTK data ────────────────────────────────────────────
for pkg in ("sentiwordnet", "wordnet", "punkt", "punkt_tab",
            "averaged_perceptron_tagger", "averaged_perceptron_tagger_eng"):
    nltk.download(pkg, quiet=True)


# ── 1. Load the Excel workbook ────────────────────────────────────────────────
EXCEL_PATH = (
    "/Users/owner/Downloads/"
    "GE Letters -database with GENDER of Sender, Recipient -editing in EXCEL.xlsx"
)

print("Loading Excel file …")
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))
headers   = rows[1]    # row index 1 = column names
data_rows = rows[2:]   # row index 2+ = data

df_raw = pd.DataFrame(data_rows, columns=headers)

title_col   = "Dublin Core:Title"
creator_col = "Dublin Core:Creator"
date_col    = "Dublin Core:Date"
text_col    = "Dublin Core:Description"


# ── 2. Filter for ALL writings authored BY George Eliot ──────────────────────
# The archive normalises her identity to "George Eliot" and "GE" abbreviations.
# Her real-name aliases (Marian Evans, Mary Ann Evans, Mary Ann Cross, etc.)
# do not appear as distinct creator values in this dataset — everything has
# already been catalogued under the GE/George Eliot umbrella.
#
# We match on:
#   • creator == "George Eliot" (letters she wrote solo)
#   • creator startswith "Ge " / "GE" — journals, diaries, co-authored items
#     (Ge Journal, Ge Diary, Ge and Ghl, Ge-To …, Ge's Draft …, etc.)
#   • title (case-insensitive) starts with "ge to " — addressed-to letters
#   • title starts with "ge and ghl to " — co-authored letters
#   • Any occurrence of her real-name aliases in creator or title
#     (belt-and-suspenders; currently zero hits, but future-proofs the script)

GE_CREATOR_EXACT = {
    "George Eliot",
    "Ge Journal", "GE Journal",
    "Ge Diary",
    "Ge and Ghl", "Ghl and Ge", "Ge And Ghl With Smith Elder & Co.",
    "Additional Letters Ge",
    "Ge's Draft For J. Chapman",
    "Ge-To Alexander Main", "Ge-To Mrs. Henry Hough",
}

GE_REAL_NAME_ALIASES = [
    "marian evans", "mary ann evans", "mary anne evans",
    "marian lewes", "marian evans lewes",
    "mary ann cross", "mary ann evans lewes",
]


def is_ge_authored(row) -> bool:
    creator = str(row[creator_col]).strip() if row[creator_col] else ""
    title   = str(row[title_col]).strip()   if row[title_col]   else ""
    c_low   = creator.lower()
    t_low   = title.lower()

    # Exact creator match
    if creator in GE_CREATOR_EXACT:
        return True
    # Title starts with "ge to" or "ge and ghl to" (case-insensitive)
    if t_low.startswith("ge to ") or t_low.startswith("ge and ghl to "):
        return True
    # Real-name aliases anywhere in creator or title
    if any(alias in c_low or alias in t_low for alias in GE_REAL_NAME_ALIASES):
        return True
    return False


mask = df_raw.apply(is_ge_authored, axis=1)
df_journals = df_raw[mask].copy()
print(f"GE-authored entries found: {len(df_journals)}")


# ── 3. Date parsing ───────────────────────────────────────────────────────────
def parse_date(raw):
    if pd.isna(raw) or not str(raw).strip() or str(raw).strip() == "None":
        return pd.NaT
    s = re.sub(r"[\s\u2060\u200b]+", " ", str(raw)).strip()
    s = re.split(r"[–—\-]{2,}|/", s)[0].strip()
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y"):
        try:
            return pd.to_datetime(s, format=fmt)
        except Exception:
            pass
    try:
        return pd.to_datetime(s, errors="coerce")
    except Exception:
        return pd.NaT


def date_from_title(title):
    months = (
        "January|February|March|April|May|June|July|"
        "August|September|October|November|December"
    )
    m = re.search(rf"(\d{{1,2}})\s+({months})\s+(\d{{4}})", str(title))
    if m:
        try:
            return pd.to_datetime(
                f"{m.group(1)} {m.group(2)} {m.group(3)}", format="%d %B %Y"
            )
        except Exception:
            pass
    m2 = re.search(rf"({months})\s+(\d{{4}})", str(title))
    if m2:
        try:
            return pd.to_datetime(f"1 {m2.group(1)} {m2.group(2)}", format="%d %B %Y")
        except Exception:
            pass
    m3 = re.search(r"\b(1\d{3})\b", str(title))
    if m3:
        try:
            return pd.to_datetime(f"1 January {m3.group(1)}", format="%d %B %Y")
        except Exception:
            pass
    return pd.NaT


df_journals["parsed_date"] = df_journals[date_col].apply(parse_date)
df_journals["parsed_date"] = df_journals.apply(
    lambda r: r["parsed_date"]
    if not pd.isna(r["parsed_date"])
    else date_from_title(r[title_col]),
    axis=1,
)

df_journals["year"]  = df_journals["parsed_date"].dt.year
df_journals["month"] = df_journals["parsed_date"].dt.month


# ── 4. Text cleaning ──────────────────────────────────────────────────────────
_SKIP = re.compile(
    r"^(GE Journal|GE to |GE and GHL|GE,|MS:|Published:|Text:|Cross,|Yale\.|"
    r"Extract published|Berg Collection|Bodleian|Pierpont Morgan|Harvard|"
    r"Beinecke|Huntington|National Library|Vol\.|vol\.|From the original)",
    re.IGNORECASE,
)

def clean_journal_text(raw_text):
    if pd.isna(raw_text) or not str(raw_text).strip():
        return ""
    text = str(raw_text).replace("_x000D_", "\n").replace("\r\n", "\n").replace("\r", "\n")
    return " ".join(
        line.strip()
        for line in text.split("\n")
        if line.strip() and not _SKIP.match(line.strip())
    )


df_journals["clean_text"] = df_journals[text_col].apply(clean_journal_text)
df_journals = df_journals[df_journals["clean_text"].str.strip().str.len() > 20].copy()
print(f"Entries with usable text: {len(df_journals)}")


# ── 5a. SentiWordNet scorer ───────────────────────────────────────────────────
# Maps Penn Treebank POS tags to WordNet POS codes so we look up the right
# synset category (noun/verb/adj/adv), improving accuracy for polysemous words.

_POS_MAP = {"J": "a", "V": "v", "N": "n", "R": "r"}


def swn_score(text: str) -> float:
    """
    Return net sentiment in [-1, 1].
    Tokenise → POS-tag → look up SentiWordNet synset for each content word
    (first / most-common synset for that POS) → mean(pos) - mean(neg).
    Returns NaN if no content words are found in the lexicon.
    """
    try:
        tokens = word_tokenize(text.lower())
        tagged = pos_tag(tokens)
    except Exception:
        return float("nan")

    pos_scores, neg_scores = [], []
    for word, tag in tagged:
        wn_pos = _POS_MAP.get(tag[0].upper())
        if wn_pos is None:
            continue
        synsets = list(swn.senti_synsets(word, wn_pos))
        if not synsets:
            continue
        s = synsets[0]          # most common sense for this POS
        # Ignore completely objective entries (both scores == 0)
        if s.pos_score() == 0 and s.neg_score() == 0:
            continue
        pos_scores.append(s.pos_score())
        neg_scores.append(s.neg_score())

    if not pos_scores:
        return float("nan")
    return float(np.mean(pos_scores) - np.mean(neg_scores))


# ── 5b. NRC Emotion Lexicon scorer ───────────────────────────────────────────
def nrc_score(text: str) -> float:
    """
    Return net sentiment in [-1, 1]:  positive_freq - negative_freq
    from the NRC Emotion Lexicon (Mohammad & Turney).
    Frequencies are normalised by the total number of affect-bearing words.
    """
    try:
        emo = NRCLex(text)
        freqs = emo.affect_frequencies
        return float(freqs.get("positive", 0.0) - freqs.get("negative", 0.0))
    except Exception:
        return float("nan")


# ── 5c. Compute scores ────────────────────────────────────────────────────────
print("Computing SentiWordNet scores …  (POS tagging takes a moment)")
df_journals["swn_score"] = df_journals["clean_text"].apply(swn_score)

print("Computing NRC Emotion Lexicon scores …")
df_journals["nrc_score"] = df_journals["clean_text"].apply(nrc_score)

# Ensemble: equal-weight average of both lexicons
df_journals["ensemble_score"] = df_journals[["swn_score", "nrc_score"]].mean(axis=1)

# Normalise to [0, 1]  (0 = most negative in corpus, 1 = most positive)
scaler = MinMaxScaler(feature_range=(0, 1))
valid_mask = df_journals["ensemble_score"].notna()
df_journals.loc[valid_mask, "norm_score"] = scaler.fit_transform(
    df_journals.loc[valid_mask, ["ensemble_score"]]
)

# ── 5d. NRC emotion breakdown (for heat-map) ──────────────────────────────────
EMOTIONS = ["joy", "sadness", "anger", "fear", "trust", "disgust", "anticipation", "surprise"]

def nrc_emotions(text: str) -> dict:
    try:
        freqs = NRCLex(text).affect_frequencies
        return {e: freqs.get(e, 0.0) for e in EMOTIONS}
    except Exception:
        return {e: 0.0 for e in EMOTIONS}

emotion_df = df_journals["clean_text"].apply(lambda t: pd.Series(nrc_emotions(t)))
df_journals = pd.concat([df_journals.reset_index(drop=True), emotion_df.reset_index(drop=True)], axis=1)


# ── 6. Aggregate by calendar month ───────────────────────────────────────────
MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun",
               "Jul","Aug","Sep","Oct","Nov","Dec"]

agg_dict = {
    "mean_swn":      ("swn_score", "mean"),
    "mean_nrc":      ("nrc_score", "mean"),
    "mean_ensemble": ("ensemble_score", "mean"),
    "mean_norm":     ("norm_score", "mean"),
    "entry_count":   ("clean_text", "count"),
}
for e in EMOTIONS:
    agg_dict[f"mean_{e}"] = (e, "mean")

monthly = (
    df_journals.groupby("month")
    .agg(**agg_dict)
    .reindex(range(1, 13))
    .reset_index()
)
monthly["month_name"] = monthly["month"].apply(lambda m: MONTH_NAMES[m - 1])
monthly["colour"] = monthly["mean_ensemble"].apply(
    lambda s: "#2ecc71" if (not pd.isna(s) and s > 0) else "#e74c3c"
)


# ── 7. Print summary table ────────────────────────────────────────────────────
print("\n══════════════════════════════════════════════════════════════════")
print("  George Eliot Journal Sentiment by Month")
print("  Lexicons: SentiWordNet (WordNet-based) + NRC Emotion Lexicon")
print("══════════════════════════════════════════════════════════════════")
print(f"  {'Month':<7} {'Entries':>7}  {'SentiWordNet':>13}  {'NRC':>7}  {'Ensemble':>9}  {'Norm(0-1)':>10}")
print("  " + "─" * 60)

for _, row in monthly.iterrows():
    if pd.isna(row["entry_count"]):
        continue
    swn_str = f"{row['mean_swn']:+.3f}" if not pd.isna(row["mean_swn"]) else "   n/a"
    nrc_str = f"{row['mean_nrc']:+.3f}" if not pd.isna(row["mean_nrc"]) else "   n/a"
    ens_str = f"{row['mean_ensemble']:+.3f}" if not pd.isna(row["mean_ensemble"]) else "   n/a"
    nrm_str = f"{row['mean_norm']:.3f}" if not pd.isna(row["mean_norm"]) else "  n/a"
    print(
        f"  {row['month_name']:<7} {int(row['entry_count']):>7}  "
        f"{swn_str:>13}  {nrc_str:>7}  {ens_str:>9}  {nrm_str:>10}"
    )

valid = monthly.dropna(subset=["mean_ensemble"])
most_pos = valid.nlargest(1,  "mean_ensemble").iloc[0]
most_neg = valid.nsmallest(1, "mean_ensemble").iloc[0]
print(f"\n  Most Positive Month : {most_pos['month_name']}  (ensemble = {most_pos['mean_ensemble']:+.3f})")
print(f"  Most Negative Month : {most_neg['month_name']}  (ensemble = {most_neg['mean_ensemble']:+.3f})")
print("══════════════════════════════════════════════════════════════════\n")


# ── 8. Visualisations ─────────────────────────────────────────────────────────
sns.set_theme(style="whitegrid", palette="muted")
fig, axes = plt.subplots(3, 1, figsize=(14, 18))
fig.suptitle(
    "George Eliot — Journal Entry Sentiment by Month\n"
    "Lexicons: SentiWordNet (WordNet-based) + NRC Emotion Lexicon\n"
    "(Only entries written BY George Eliot)",
    fontsize=13, fontweight="bold", y=0.995,
)

valid_m = monthly.dropna(subset=["mean_ensemble"])

# ── Plot 1: Ensemble bar chart ────────────────────────────────────────────────
ax1 = axes[0]
bars = ax1.bar(
    valid_m["month_name"], valid_m["mean_ensemble"],
    color=valid_m["colour"], edgecolor="white", linewidth=0.8, zorder=3,
)
ax1.axhline(0, color="black", linewidth=0.9, linestyle="--", alpha=0.6)
ax1.set_title("Ensemble Sentiment Score by Month (SentiWordNet + NRC avg)",
              fontsize=11, fontweight="bold")
ax1.set_ylabel("Mean Score  (−1 = very negative,  +1 = very positive)")
ax1.set_xlabel("Month")
ax1.set_ylim(-0.5, 0.5)
ax1.grid(axis="y", alpha=0.4, zorder=0)

for bar, val in zip(bars, valid_m["mean_ensemble"]):
    if not pd.isna(val):
        offset = 0.012 if val >= 0 else -0.025
        va     = "bottom" if val >= 0 else "top"
        ax1.text(bar.get_x() + bar.get_width() / 2, val + offset,
                 f"{val:+.2f}", ha="center", va=va, fontsize=8)

pos_p = mpatches.Patch(color="#2ecc71", label="Net positive")
neg_p = mpatches.Patch(color="#e74c3c", label="Net negative")
ax1.legend(handles=[pos_p, neg_p], loc="upper right", fontsize=9)

# ── Plot 2: SentiWordNet vs NRC side-by-side ──────────────────────────────────
ax2 = axes[1]
x = np.arange(len(valid_m))
w = 0.35
ax2.bar(x - w/2, valid_m["mean_swn"], w, label="SentiWordNet",       color="#2980b9", alpha=0.85, zorder=3)
ax2.bar(x + w/2, valid_m["mean_nrc"], w, label="NRC Emotion Lexicon", color="#8e44ad", alpha=0.85, zorder=3)
ax2.axhline(0, color="black", linewidth=0.9, linestyle="--", alpha=0.6)
ax2.set_title("SentiWordNet vs. NRC Emotion Lexicon Scores by Month",
              fontsize=11, fontweight="bold")
ax2.set_ylabel("Mean Polarity Score")
ax2.set_xlabel("Month")
ax2.set_xticks(x)
ax2.set_xticklabels(valid_m["month_name"])
ax2.legend(fontsize=9)
ax2.grid(axis="y", alpha=0.4, zorder=0)

# ── Plot 3: NRC emotion heat-map by month ─────────────────────────────────────
ax3 = axes[2]
heat_data = valid_m.set_index("month_name")[[f"mean_{e}" for e in EMOTIONS]].copy()
heat_data.columns = [e.capitalize() for e in EMOTIONS]
heat_data = heat_data.fillna(0).T    # emotions as rows, months as columns

sns.heatmap(
    heat_data, ax=ax3,
    cmap="RdYlGn", center=0,
    linewidths=0.5, linecolor="white",
    annot=True, fmt=".2f", annot_kws={"size": 8},
    cbar_kws={"label": "NRC frequency (proportion of affect words)"},
)
ax3.set_title("NRC Emotion Profile by Month — George Eliot Journal Entries",
              fontsize=11, fontweight="bold")
ax3.set_xlabel("Month")
ax3.set_ylabel("Emotion")
ax3.tick_params(axis="x", rotation=0)
ax3.tick_params(axis="y", rotation=0)

plt.tight_layout(rect=[0, 0, 1, 0.975])
out_png = "/Users/owner/Downloads/ge_journal_sentiment.png"
plt.savefig(out_png, dpi=150, bbox_inches="tight")
print(f"Chart saved → {out_png}")
plt.show()


# ── 9. Save results ───────────────────────────────────────────────────────────
save_cols = [
    title_col, date_col, "parsed_date", "year", "month",
    "swn_score", "nrc_score", "ensemble_score", "norm_score", "clean_text",
]
df_journals[save_cols].to_csv(
    "/Users/owner/Downloads/ge_journal_sentiment_scores.csv", index=False
)
monthly.to_csv(
    "/Users/owner/Downloads/ge_journal_monthly_summary.csv", index=False
)
print("Per-entry scores → ge_journal_sentiment_scores.csv")
print("Monthly summary  → ge_journal_monthly_summary.csv")


# ── 10. Print HTML-ready monthStats for george-eliot-visualizations.html ──────
# Computes raw NRC positive/negative word counts and average word count so the
# Seasonal Mood Wheel in the HTML can be updated with the full-corpus results.

MONTH_FULL = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]

def nrc_raw_counts(text: str):
    """Return (pos_word_count, neg_word_count) from the NRC Emotion Lexicon."""
    try:
        scores = NRCLex(text).raw_emotion_scores
        return scores.get("positive", 0), scores.get("negative", 0)
    except Exception:
        return 0, 0

print("\nComputing raw NRC word counts for HTML output …")
df_journals[["nrc_pos_raw", "nrc_neg_raw"]] = df_journals["clean_text"].apply(
    lambda t: pd.Series(nrc_raw_counts(t))
)
df_journals["word_count"] = df_journals["clean_text"].apply(
    lambda t: len(str(t).split())
)

html_monthly = (
    df_journals.groupby("month")
    .agg(
        letters   =("clean_text",    "count"),
        pos       =("nrc_pos_raw",   "sum"),
        neg       =("nrc_neg_raw",   "sum"),
        avg_words =("word_count",    "mean"),
    )
    .reindex(range(1, 13), fill_value=0)
    .reset_index()
)

print("\n// ── Paste this into george-eliot-visualizations.html (monthStats) ──")
print("const monthStats = [")
for _, r in html_monthly.iterrows():
    m      = int(r["month"])
    name   = MONTH_FULL[m - 1]
    lets   = int(r["letters"])
    pos    = int(r["pos"])
    neg    = int(r["neg"])
    avg_w  = int(round(r["avg_words"]))
    print(f"  {{ month:{m:2d}, name:'{name}',{'':<3} letters:{lets}, pos:{pos}, neg:{neg}, avgWords:{avg_w} }},")
print("];")
