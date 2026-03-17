#!/usr/bin/env python3
"""
process_data.py — Extract correspondence data from George Eliot Letters XLSX
and produce a compact JSON for the interactive visualization.

Usage:
    python3 process_data.py <path_to_xlsx> <output_json>

Example:
    python3 scripts/process_data.py \
        "../Letters/GE Letters -database with GENDER of Sender, Recipient -editing in EXCEL.xlsx" \
        data/letters.json
"""

import json
import sys
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def extract_data(xlsx_path):
    """Read XLSX and return structured letter data."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    letters = []
    total = 0

    for i, row in enumerate(ws.iter_rows(min_row=3)):  # Row 1 = title, Row 2 = headers
        vals = [cell.value for cell in row]

        item_id = vals[0]
        if not item_id:
            continue

        recip_gender = str(vals[3]).strip().upper() if vals[3] else "U"
        sender_gender = str(vals[4]).strip().upper() if vals[4] else "U"
        sender_name = str(vals[7]).strip() if vals[7] else "Unknown"
        recip_name = str(vals[5]).strip() if vals[5] else "Unknown"
        date_str = str(vals[10]).strip() if vals[10] else ""

        # Normalize genders
        if sender_gender not in ("M", "F"):
            sender_gender = "U"
        if recip_gender not in ("M", "F"):
            recip_gender = "U"

        # Extract year
        year = None
        if date_str and len(date_str) >= 4:
            try:
                year = int(date_str[:4])
            except ValueError:
                pass

        total += 1
        letters.append({
            "year": year,
            "sg": sender_gender,
            "rg": recip_gender,
            "sender": sender_name,
            "recipient": recip_name,
        })

    wb.close()
    return letters, total


def compute_aggregates(letters):
    """Compute matrices, decade breakdowns, yearly counts, top correspondents."""

    # Gender pair matrix (only M/F)
    matrix = [[0, 0], [0, 0]]  # [sender_idx][recip_idx], 0=F, 1=M
    gender_idx = {"F": 0, "M": 1}

    by_year = defaultdict(lambda: {"FF": 0, "FM": 0, "MF": 0, "MM": 0})
    by_decade = defaultdict(lambda: {"FF": 0, "FM": 0, "MF": 0, "MM": 0})

    sender_stats = defaultdict(lambda: {"gender": "U", "sent": 0, "received": 0})
    recip_stats = defaultdict(lambda: {"gender": "U", "sent": 0, "received": 0})

    for letter in letters:
        sg, rg = letter["sg"], letter["rg"]

        # Update sender/recipient stats
        sender_stats[letter["sender"]]["sent"] += 1
        sender_stats[letter["sender"]]["gender"] = sg
        recip_stats[letter["recipient"]]["received"] += 1
        recip_stats[letter["recipient"]]["gender"] = rg

        if sg in gender_idx and rg in gender_idx:
            si, ri = gender_idx[sg], gender_idx[rg]
            matrix[si][ri] += 1

            pair_key = sg + rg
            if letter["year"]:
                by_year[str(letter["year"])][pair_key] += 1
                decade = (letter["year"] // 10) * 10
                decade_label = f"{decade}s"
                by_decade[decade_label][pair_key] += 1

    # Merge sender/recipient stats into top correspondents
    all_people = set(list(sender_stats.keys()) + list(recip_stats.keys()))
    people = []
    for name in all_people:
        sent = sender_stats[name]["sent"]
        received = recip_stats[name]["received"]
        gender = sender_stats[name]["gender"]
        if gender == "U":
            gender = recip_stats[name]["gender"]
        if sent + received >= 10:  # Only include people with 10+ letters
            people.append({
                "name": name,
                "gender": gender,
                "sent": sent,
                "received": received,
                "total": sent + received,
            })

    people.sort(key=lambda x: x["total"], reverse=True)

    # Sort by_year keys
    by_year_sorted = {k: by_year[k] for k in sorted(by_year.keys())}
    by_decade_sorted = {k: by_decade[k] for k in sorted(by_decade.keys())}

    return matrix, by_year_sorted, by_decade_sorted, people[:30]


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_path = sys.argv[2]

    print(f"Reading {xlsx_path} ...")
    letters, total = extract_data(xlsx_path)
    print(f"  {total} letters extracted")

    gendered = [l for l in letters if l["sg"] in ("M", "F") and l["rg"] in ("M", "F")]
    print(f"  {len(gendered)} with identified sender & recipient genders")

    matrix, by_year, by_decade, top_correspondents = compute_aggregates(letters)

    # Build compact letter list (only gendered, with year)
    compact_letters = []
    for l in letters:
        if l["year"] and l["sg"] in ("M", "F") and l["rg"] in ("M", "F"):
            compact_letters.append({
                "y": l["year"],
                "sg": l["sg"],
                "rg": l["rg"],
                "s": l["sender"],
                "r": l["recipient"],
            })

    years_with_data = [l["y"] for l in compact_letters]
    date_range = [min(years_with_data), max(years_with_data)] if years_with_data else [0, 0]

    output = {
        "metadata": {
            "totalLetters": total,
            "genderedLetters": len(gendered),
            "dateRange": date_range,
        },
        "matrix": matrix,
        "byYear": by_year,
        "byDecade": by_decade,
        "topCorrespondents": top_correspondents,
        "letters": compact_letters,
    }

    with open(output_path, "w") as f:
        json.dump(output, f, separators=(",", ":"))

    size_kb = len(json.dumps(output, separators=(",", ":"))) / 1024
    print(f"  Output: {output_path} ({size_kb:.0f} KB)")
    print("Done.")


if __name__ == "__main__":
    main()
